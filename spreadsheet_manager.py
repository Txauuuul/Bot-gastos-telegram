"""
Módulo para gestionar Excel con gastos.
Soporta almacenamiento local y en Google Drive.
Incluye análisis, reportes, búsqueda y gestión inteligente.
"""

import os
import json
import logging
from io import BytesIO
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
except ImportError:
    pass

logger = logging.getLogger(__name__)

# ID de carpeta en Google Drive donde guardar los archivos (opcional)
GOOGLE_DRIVE_FOLDER_ID = None


class SpreadsheetManager:
    """Gestiona hojas de cálculo con gastos avanzadas."""

    LOCAL_FILE = "gastos.xlsx"
    CONFIG_FILE = "config.json"
    CATEGORIES_FILE = "categorias.json"
    DELETED_HISTORY_FILE = "historial_borrados.json"
    SCOPES = ["https://www.googleapis.com/auth/drive.file"]

    def __init__(self, use_google_drive: bool = False):
        """
        Inicializa el gestor de hojas de cálculo.

        Args:
            use_google_drive: Si True, sincroniza con Google Drive
        """
        self.use_google_drive = use_google_drive
        self.drive_service = None
        self.drive_file_id = None
        self.config = self._cargar_config()
        self.categorias_personalizadas = self._cargar_categorias()

        if use_google_drive:
            self._init_google_drive()
            # Descargar Excel desde Google Drive al iniciar (recupera gastos tras redeploy)
            self._descargar_excel_del_google_drive()

    def _cargar_config(self) -> Dict:
        """Carga la configuración desde archivo JSON."""
        try:
            if os.path.exists(self.CONFIG_FILE):
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.warning(f"No se pudo cargar configuración: {e}")
        
        return {
            "presupuesto_diario": 100,
            "presupuesto_mensual": 3000,
            "dias_sin_gastos_alerta": 7,
            "recordatorio_viernes": True,
            "recordatorio_fin_mes": True,
            "ultima_sincronizacion": None
        }
    
    def _guardar_config(self) -> bool:
        """Guarda la configuración en archivo JSON."""
        try:
            with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f)
            return True
        except Exception as e:
            logger.error(f"Error guardando configuración: {e}")
            return False
    
    def _cargar_categorias(self) -> List[str]:
        """Carga las categorías personalizadas."""
        try:
            if os.path.exists(self.CATEGORIES_FILE):
                with open(self.CATEGORIES_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.warning(f"No se pudo cargar categorías: {e}")
        
        return ["Supervivencia", "Electrónicos", "Viajes", "Caprichos"]
    
    def _guardar_categorias(self) -> bool:
        """Guarda las categorías en archivo JSON."""
        try:
            with open(self.CATEGORIES_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.categorias_personalizadas, f)
            return True
        except Exception as e:
            logger.error(f"Error guardando categorías: {e}")
            return False
    
    def _cargar_historial_borrados(self) -> List[Dict]:
        """Carga el historial de gastos borrados."""
        try:
            if os.path.exists(self.DELETED_HISTORY_FILE):
                with open(self.DELETED_HISTORY_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
        except Exception as e:
            logger.warning(f"No se pudo cargar historial: {e}")
        return []
    
    def _guardar_historial_borrados(self, historial: List[Dict]) -> bool:
        """Guarda el historial de gastos borrados."""
        try:
            with open(self.DELETED_HISTORY_FILE, 'w', encoding='utf-8') as f:
                json.dump(historial, f)
            return True
        except Exception as e:
            logger.error(f"Error guardando historial: {e}")
            return False

    def _init_google_drive(self) -> bool:
        """Inicializa autenticación con Google Drive."""
        try:
            if os.path.exists("token.json"):
                creds = Credentials.from_authorized_user_file("token.json", self.SCOPES)
            else:
                if not os.path.exists("credentials.json"):
                    logger.warning(
                        "⚠️  credentials.json no encontrado. "
                        "Google Drive deshabilitado. "
                        "Guarda en local solamente."
                    )
                    self.use_google_drive = False
                    return False

                flow = InstalledAppFlow.from_client_secrets_file(
                    "credentials.json", self.SCOPES
                )
                creds = flow.run_local_server(port=0)

                with open("token.json", "w") as token:
                    token.write(creds.to_json())

            self.drive_service = build("drive", "v3", credentials=creds)
            logger.info("✅ Google Drive autenticado")
            return True

        except Exception as e:
            logger.error(f"❌ Error en Google Drive: {e}")
            self.use_google_drive = False
            return False

    def _crear_excel_base(self) -> None:
        """Crea un Excel nuevo con encabezados."""
        if not openpyxl:
            raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gastos"

        encabezados = ["Fecha", "Hora", "Concepto", "Precio (€)", "Categoría"]
        ws.append(encabezados)

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 15

        wb.save(self.LOCAL_FILE)
        logger.info(f"📄 Archivo {self.LOCAL_FILE} creado")

    def agregar_gasto(self, gasto: Dict) -> bool:
        """
        Agrega un gasto al Excel.

        Args:
            gasto: Dict con concepto, precio, fecha, hora

        Returns:
            True si se agregó exitosamente
        """
        try:
            if not os.path.exists(self.LOCAL_FILE):
                self._crear_excel_base()

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            fila_siguiente = ws.max_row + 1

            ws[f"A{fila_siguiente}"] = gasto.get("fecha", "")
            ws[f"B{fila_siguiente}"] = gasto.get("hora", "")
            ws[f"C{fila_siguiente}"] = gasto.get("concepto", "")
            ws[f"D{fila_siguiente}"] = gasto.get("precio", 0)
            ws[f"E{fila_siguiente}"] = gasto.get("categoria", "General")

            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col in ["A", "B", "C", "D", "E"]:
                celda = ws[f"{col}{fila_siguiente}"]
                celda.border = border
                celda.alignment = Alignment(horizontal="center" if col != "C" else "left")

                if col == "D":
                    celda.number_format = "0.00"

            wb.save(self.LOCAL_FILE)
            logger.info(f"✅ Gasto agregado: {gasto['concepto']} - {gasto['precio']}€")

            if self.use_google_drive and self.drive_service:
                self._sincronizar_google_drive()

            return True

        except Exception as e:
            logger.error(f"❌ Error al agregar gasto: {e}")
            return False

    def _descargar_excel_del_google_drive(self) -> bool:
        """Descarga el Excel desde Google Drive al iniciar (recupera gastos tras redeploy)."""
        try:
            if not self.drive_service:
                logger.warning("Google Drive no está configurado para descargar")
                return False

            query = f"name='{self.LOCAL_FILE}' and trashed=false"
            results = self.drive_service.files().list(q=query, spaces="drive", pageSize=1).execute()
            files = results.get("files", [])

            if not files:
                logger.info(f"📄 No hay {self.LOCAL_FILE} en Google Drive (primera vez)")
                return False

            file_id = files[0]["id"]
            request = self.drive_service.files().get_media(fileId=file_id)
            
            fh = BytesIO()
            downloader = MediaIoBaseDownload(fh, request)

            done = False
            while not done:
                status, done = downloader.next_chunk()

            fh.seek(0)
            with open(self.LOCAL_FILE, 'wb') as f:
                f.write(fh.read())

            logger.info(f"☁️  Excel descargado desde Google Drive (recuperado tras redeploy)")
            return True

        except Exception as e:
            logger.error(f"❌ Error al descargar desde Google Drive: {e}")
            return False

    def _sincronizar_google_drive(self) -> bool:
        """Carga el Excel a Google Drive INMEDIATAMENTE (sin esperar 12 horas)."""
        try:
            if not self.drive_service:
                logger.warning("Google Drive no está configurado")
                return False

            query = f"name='{self.LOCAL_FILE}' and trashed=false"
            results = self.drive_service.files().list(q=query, spaces="drive", pageSize=1).execute()
            files = results.get("files", [])

            if files:
                file_id = files[0]["id"]
                media = MediaFileUpload(self.LOCAL_FILE, resumable=True)
                self.drive_service.files().update(fileId=file_id, media_body=media).execute()
                logger.info(f"☁️  Excel sincronizado con Google Drive (gasto guardado)")
            else:
                file_metadata = {"name": self.LOCAL_FILE}
                media = MediaFileUpload(self.LOCAL_FILE, resumable=True)
                file = self.drive_service.files().create(body=file_metadata, media_body=media).execute()
                logger.info(f"☁️  Excel subido a Google Drive ({file['id']})")

            return True

        except Exception as e:
            logger.error(f"❌ Error al sincronizar con Google Drive: {e}")
            return False

    def obtener_resumen(self) -> Dict:
        """Obtiene un resumen de gastos del mes actual."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {"total": 0, "cantidad": 0, "promedio": 0}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            total = 0
            cantidad = 0
            mes_actual = datetime.now().month

            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[0]:
                    try:
                        fecha = datetime.strptime(str(fila[0]), "%d/%m/%Y")
                        if fecha.month == mes_actual:
                            precio = fila[3] or 0
                            total += float(precio)
                            cantidad += 1
                    except:
                        pass

            promedio = total / cantidad if cantidad > 0 else 0

            return {
                "total": round(total, 2),
                "cantidad": cantidad,
                "promedio": round(promedio, 2),
            }

        except Exception as e:
            logger.error(f"❌ Error al obtener resumen: {e}")
            return {"total": 0, "cantidad": 0, "promedio": 0}

    def obtener_gastos_ultimos_dias(self, dias: int) -> Dict:
        """Obtiene todos los gastos de los últimos X días."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {
                    "gastos": [],
                    "total": 0,
                    "cantidad": 0,
                    "mensaje": "❌ No hay gastos registrados"
                }

            if dias <= 0:
                return {
                    "gastos": [],
                    "total": 0,
                    "cantidad": 0,
                    "mensaje": "❌ El número de días debe ser mayor a 0"
                }

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos = []
            total = 0
            fecha_limite = datetime.now().timestamp() - (dias * 86400)

            for fila in list(ws.iter_rows(min_row=2, values_only=False)):
                if fila[0].value:
                    try:
                        fecha_str = str(fila[0].value)
                        fecha_obj = datetime.strptime(fecha_str, "%d/%m/%Y")
                        
                        if fecha_obj.timestamp() >= fecha_limite:
                            precio = float(fila[3].value or 0)
                            gasto = {
                                "fecha": fila[0].value,
                                "hora": fila[1].value or "",
                                "concepto": fila[2].value or "",
                                "precio": precio,
                                "categoria": fila[4].value or "General"
                            }
                            gastos.append(gasto)
                            total += precio
                    except (ValueError, TypeError):
                        pass

            gastos.sort(key=lambda x: x["hora"] if x["hora"] else "00:00:00", reverse=True)
            gastos.sort(key=lambda x: x["fecha"], reverse=True)

            return {
                "gastos": gastos,
                "total": round(total, 2),
                "cantidad": len(gastos),
                "dias": dias
            }

        except Exception as e:
            logger.error(f"❌ Error al obtener gastos: {e}")
            return {
                "gastos": [],
                "total": 0,
                "cantidad": 0,
                "mensaje": f"❌ Error: {str(e)}"
            }

    def eliminar_gasto(self, concepto: str) -> Dict:
        """Elimina el último gasto con ese concepto y guarda en historial."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {"exito": False, "mensaje": "❌ No hay gastos registrados"}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            filas_data = list(ws.iter_rows(min_row=2))
            fila_encontrada = None
            fila_numero = None

            for i in range(len(filas_data) - 1, -1, -1):
                fila = filas_data[i]
                if fila[2].value and fila[2].value.lower().strip() == concepto.lower().strip():
                    fila_encontrada = fila
                    fila_numero = i + 2
                    break

            if not fila_encontrada:
                return {
                    "exito": False,
                    "mensaje": f"❌ No encontré ningún gasto con concepto '{concepto}'"
                }

            gasto_borrado = {
                "fecha": str(fila_encontrada[0].value) if fila_encontrada[0].value else "",
                "hora": str(fila_encontrada[1].value) if fila_encontrada[1].value else "",
                "concepto": fila_encontrada[2].value or "",
                "precio": float(fila_encontrada[3].value or 0),
                "categoria": fila_encontrada[4].value or "General",
                "fecha_borrado": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }

            historial = self._cargar_historial_borrados()
            historial.append(gasto_borrado)
            self._guardar_historial_borrados(historial)

            ws.delete_rows(fila_numero, 1)

            wb.save(self.LOCAL_FILE)
            logger.info(f"✅ Gasto eliminado: {gasto_borrado['concepto']} - {gasto_borrado['precio']}€")

            if self.use_google_drive and self.drive_service:
                self._sincronizar_google_drive()

            return {
                "exito": True,
                "mensaje": f"✅ Gasto eliminado: {gasto_borrado['concepto']} ({gasto_borrado['precio']}€)",
                "concepto": gasto_borrado['concepto'],
                "precio": gasto_borrado['precio']
            }

        except Exception as e:
            logger.error(f"❌ Error al eliminar gasto: {e}")
            return {"exito": False, "mensaje": f"❌ Error al eliminar: {str(e)}"}

    def deshacer_ultimo_gasto(self) -> Dict:
        """Restaura el último gasto borrado."""
        try:
            historial = self._cargar_historial_borrados()
            if not historial:
                return {"exito": False, "mensaje": "❌ No hay gastos borrados para restaurar"}
            
            gasto_restaurado = historial.pop()
            
            if self.agregar_gasto({
                "fecha": gasto_restaurado["fecha"],
                "hora": gasto_restaurado["hora"],
                "concepto": gasto_restaurado["concepto"],
                "precio": gasto_restaurado["precio"],
                "categoria": gasto_restaurado["categoria"]
            }):
                self._guardar_historial_borrados(historial)
                return {
                    "exito": True,
                    "mensaje": f"✅ Gasto restaurado: {gasto_restaurado['concepto']} ({gasto_restaurado['precio']}€)"
                }
            return {"exito": False, "mensaje": "❌ Error al restaurar"}
        except Exception as e:
            logger.error(f"Error deshaciendo: {e}")
            return {"exito": False, "mensaje": f"❌ Error: {str(e)}"}

    def obtener_gastos_por_categoria(self) -> Dict:
        """Retorna total de gastos por categoría del mes actual."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            categorias = {}
            mes_actual = datetime.now().month

            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[0]:
                    try:
                        fecha = datetime.strptime(str(fila[0]), "%d/%m/%Y")
                        if fecha.month == mes_actual:
                            categoria = fila[4] or "General"
                            precio = float(fila[3] or 0)
                            categorias[categoria] = categorias.get(categoria, 0) + precio
                    except:
                        pass

            return {k: round(v, 2) for k, v in sorted(categorias.items(), key=lambda x: x[1], reverse=True)}
        except Exception as e:
            logger.error(f"Error obteniendo gastos por categoría: {e}")
            return {}

    def obtener_promedio_diario(self) -> float:
        """Calcula el promedio de gasto diario del mes actual."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return 0

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos_por_dias = {}
            mes_actual = datetime.now().month
            
            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[0]:
                    try:
                        fecha_str = str(fila[0])
                        fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
                        
                        if fecha.month == mes_actual:
                            precio = float(fila[3] or 0)
                            
                            if fecha_str not in gastos_por_dias:
                                gastos_por_dias[fecha_str] = 0
                            gastos_por_dias[fecha_str] += precio
                    except:
                        pass
            
            if not gastos_por_dias:
                return 0
            
            total = sum(gastos_por_dias.values())
            return round(total / len(gastos_por_dias), 2)
        except Exception as e:
            logger.error(f"Error calculando promedio: {e}")
            return 0

    def obtener_gasto_del_dia(self, fecha_str: str = None) -> Dict:
        """Obtiene los gastos de un día específico."""
        if fecha_str is None:
            fecha_str = datetime.now().strftime("%d/%m/%Y")
        
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {"gastos": [], "total": 0}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos = []
            total = 0

            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[0] and str(fila[0]) == fecha_str:
                    precio = float(fila[3] or 0)
                    gastos.append({
                        "concepto": fila[2],
                        "precio": precio,
                        "categoria": fila[4] or "General"
                    })
                    total += precio

            return {"gastos": gastos, "total": round(total, 2), "fecha": fecha_str}
        except Exception as e:
            logger.error(f"Error obteniendo gasto del día: {e}")
            return {"gastos": [], "total": 0}

    def obtener_gastos_por_mes(self, mes: int, año: int = None) -> Dict:
        """Obtiene gastos de un mes específico."""
        if año is None:
            año = datetime.now().year
        
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {"gastos": [], "total": 0, "cantidad": 0}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos = []
            total = 0

            for fila in list(ws.iter_rows(min_row=2, values_only=False)):
                if fila[0].value:
                    try:
                        fecha = datetime.strptime(str(fila[0].value), "%d/%m/%Y")
                        if fecha.month == mes and fecha.year == año:
                            precio = float(fila[3].value or 0)
                            gastos.append({
                                "fecha": str(fila[0].value),
                                "concepto": fila[2].value,
                                "precio": precio,
                                "categoria": fila[4].value or "General"
                            })
                            total += precio
                    except:
                        pass

            return {
                "gastos": sorted(gastos, key=lambda x: x["fecha"], reverse=True),
                "total": round(total, 2),
                "cantidad": len(gastos)
            }
        except Exception as e:
            logger.error(f"Error obteniendo gastos del mes: {e}")
            return {"gastos": [], "total": 0, "cantidad": 0}

    def obtener_gastos_entre_fechas(self, fecha_inicio: str, fecha_fin: str) -> Dict:
        """Obtiene gastos entre dos fechas (formato DD/MM/YYYY)."""
        try:
            fecha_ini = datetime.strptime(fecha_inicio, "%d/%m/%Y")
            fecha_fin_obj = datetime.strptime(fecha_fin, "%d/%m/%Y")
            
            if not os.path.exists(self.LOCAL_FILE):
                return {"gastos": [], "total": 0, "cantidad": 0}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos = []
            total = 0

            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[0]:
                    try:
                        fecha = datetime.strptime(str(fila[0]), "%d/%m/%Y")
                        if fecha_ini <= fecha <= fecha_fin_obj:
                            precio = float(fila[3] or 0)
                            gastos.append({
                                "fecha": str(fila[0]),
                                "hora": fila[1] or "",
                                "concepto": fila[2],
                                "precio": precio,
                                "categoria": fila[4] or "General"
                            })
                            total += precio
                    except:
                        pass

            return {
                "gastos": sorted(gastos, key=lambda x: x["fecha"], reverse=True),
                "total": round(total, 2),
                "cantidad": len(gastos),
                "periodo": f"{fecha_inicio} a {fecha_fin}"
            }
        except ValueError:
            return {"gastos": [], "total": 0, "cantidad": 0, "error": "Formato incorrecto (usar DD/MM/YYYY)"}

    def buscar_por_concepto(self, palabra_clave: str) -> Dict:
        """Busca gastos por palabra clave."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {"gastos": [], "total": 0}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos = []
            total = 0
            palabra_lower = palabra_clave.lower()

            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[2] and palabra_lower in str(fila[2]).lower():
                    precio = float(fila[3] or 0)
                    gastos.append({
                        "fecha": fila[0],
                        "hora": fila[1] or "",
                        "concepto": fila[2],
                        "precio": precio,
                        "categoria": fila[4] or "General"
                    })
                    total += precio

            return {
                "gastos": sorted(gastos, key=lambda x: x["fecha"], reverse=True),
                "total": round(total, 2),
                "cantidad": len(gastos),
                "palabra_clave": palabra_clave
            }
        except Exception as e:
            logger.error(f"Error buscando: {e}")
            return {"gastos": [], "total": 0}

    def obtener_top_gastos(self, cantidad: int = 5) -> List[Dict]:
        """Retorna los gastos más caros."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return []

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos = []
            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[0]:
                    precio = float(fila[3] or 0)
                    gastos.append({
                        "fecha": fila[0],
                        "concepto": fila[2],
                        "precio": precio,
                        "categoria": fila[4] or "General"
                    })

            return sorted(gastos, key=lambda x: x["precio"], reverse=True)[:cantidad]
        except Exception as e:
            logger.error(f"Error obteniendo top gastos: {e}")
            return []

    def obtener_ranking_gastos(self) -> Dict:
        """Retorna estadísticas de conceptos más frecuentes."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            conceptos = {}
            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[2]:
                    concepto = fila[2]
                    precio = float(fila[3] or 0)
                    if concepto not in conceptos:
                        conceptos[concepto] = {"total": 0, "cantidad": 0}
                    conceptos[concepto]["total"] += precio
                    conceptos[concepto]["cantidad"] += 1

            ranking = {}
            for concepto, datos in conceptos.items():
                ranking[concepto] = {
                    "total": round(datos["total"], 2),
                    "cantidad": datos["cantidad"],
                    "promedio": round(datos["total"] / datos["cantidad"], 2)
                }
            
            return dict(sorted(ranking.items(), key=lambda x: x[1]["total"], reverse=True))
        except Exception as e:
            logger.error(f"Error en ranking: {e}")
            return {}

    def obtener_proyeccion_mes(self) -> Dict:
        """Proyecta el gasto total del mes actual."""
        try:
            hoy = datetime.now()
            dias_transcurridos = hoy.day
            
            resultado_mes = self.obtener_gastos_por_mes(hoy.month, hoy.year)
            gasto_actual = resultado_mes["total"]
            
            if dias_transcurridos == 0:
                return {"error": "Aún no hay datos del día"}
            
            promedio_diario = gasto_actual / dias_transcurridos
            dias_totales_mes = 30 if hoy.month in [4, 6, 9, 11] else 31 if hoy.month != 2 else 28
            proyeccion = promedio_diario * dias_totales_mes
            
            return {
                "gasto_actual": round(gasto_actual, 2),
                "dias_transcurridos": dias_transcurridos,
                "promedio_diario": round(promedio_diario, 2),
                "proyeccion_mes": round(proyeccion, 2),
                "diferencia_presupuesto": round(self.config.get("presupuesto_mensual", 3000) - proyeccion, 2)
            }
        except Exception as e:
            logger.error(f"Error en proyección: {e}")
            return {}

    def obtener_ahorro_potencial(self) -> Dict:
        """Calcula cuánto se podría ahorrar reduciendo categorías."""
        try:
            categorias_gasto = self.obtener_gastos_por_categoria()
            ahorro = {}
            
            for categoria, total in categorias_gasto.items():
                ahorro[categoria] = {
                    "gasto_actual": total,
                    "reducir_10%": round(total * 0.1, 2),
                    "reducir_25%": round(total * 0.25, 2),
                    "reducir_50%": round(total * 0.5, 2)
                }
            
            return ahorro
        except Exception as e:
            logger.error(f"Error calculando ahorro: {e}")
            return {}

    def obtener_historial_gastos(self, cantidad: int = 5) -> List[Dict]:
        """Retorna los últimos X gastos registrados."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return []

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            gastos = []
            for fila in list(ws.iter_rows(min_row=2, values_only=True)):
                if fila[0]:
                    gastos.append({
                        "fecha": fila[0],
                        "hora": fila[1] or "",
                        "concepto": fila[2],
                        "precio": float(fila[3] or 0),
                        "categoria": fila[4] or "General"
                    })

            return sorted(gastos, key=lambda x: (str(x["fecha"]), x["hora"]), reverse=True)[:cantidad]
        except Exception as e:
            logger.error(f"Error obteniendo historial: {e}")
            return []

    def editargasto(self, concepto_original: str, nuevo_precio: float = None, nueva_categoria: str = None) -> Dict:
        """Edita el precio o categoría del último gasto."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {"exito": False, "mensaje": "❌ No hay gastos registrados"}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            filas_data = list(ws.iter_rows(min_row=2))
            fila_encontrada_idx = None

            for i in range(len(filas_data) - 1, -1, -1):
                fila = filas_data[i]
                if fila[2].value and fila[2].value.lower().strip() == concepto_original.lower().strip():
                    fila_encontrada_idx = i
                    break

            if fila_encontrada_idx is None:
                return {"exito": False, "mensaje": f"❌ No encontré gasto: {concepto_original}"}

            fila_numero = fila_encontrada_idx + 2

            if nuevo_precio is not None:
                ws[f"D{fila_numero}"] = nuevo_precio
            if nueva_categoria is not None:
                ws[f"E{fila_numero}"] = nueva_categoria

            wb.save(self.LOCAL_FILE)
            if self.use_google_drive and self.drive_service:
                self._sincronizar_google_drive()

            msg = f"✅ Gasto '{concepto_original}' actualizado"
            if nuevo_precio:
                msg += f" (precio: {nuevo_precio}€)"
            if nueva_categoria:
                msg += f" (categoría: {nueva_categoria})"
            
            return {"exito": True, "mensaje": msg}
        except Exception as e:
            logger.error(f"Error editando gasto: {e}")
            return {"exito": False, "mensaje": f"❌ Error: {str(e)}"}

    def duplicar_gasto(self, concepto: str) -> Dict:
        """Agrega el mismo gasto nuevamente."""
        try:
            if not os.path.exists(self.LOCAL_FILE):
                return {"exito": False, "mensaje": "❌ No hay gastos registrados"}

            wb = openpyxl.load_workbook(self.LOCAL_FILE)
            ws = wb.active

            filas_data = list(ws.iter_rows(min_row=2))
            fila_encontrada = None

            for i in range(len(filas_data) - 1, -1, -1):
                fila = filas_data[i]
                if fila[2].value and fila[2].value.lower().strip() == concepto.lower().strip():
                    fila_encontrada = fila
                    break

            if not fila_encontrada:
                return {"exito": False, "mensaje": f"❌ No encontré gasto: {concepto}"}

            nuevo_gasto = {
                "fecha": datetime.now().strftime("%d/%m/%Y"),
                "hora": datetime.now().strftime("%H:%M:%S"),
                "concepto": fila_encontrada[2].value,
                "precio": float(fila_encontrada[3].value or 0),
                "categoria": fila_encontrada[4].value or "General"
            }

            if self.agregar_gasto(nuevo_gasto):
                return {
                    "exito": True,
                    "mensaje": f"✅ Gasto duplicado: {nuevo_gasto['concepto']} ({nuevo_gasto['precio']}€)"
                }
            return {"exito": False, "mensaje": "❌ Error al duplicar"}
        except Exception as e:
            logger.error(f"Error duplicando gasto: {e}")
            return {"exito": False, "mensaje": f"❌ Error: {str(e)}"}

    def agregar_categoria(self, categoria: str) -> Dict:
        """Agrega una nueva categoría personalizada."""
        if categoria in self.categorias_personalizadas:
            return {"exito": False, "mensaje": f"❌ La categoría '{categoria}' ya existe"}
        
        self.categorias_personalizadas.append(categoria)
        if self._guardar_categorias():
            return {"exito": True, "mensaje": f"✅ Categoría '{categoria}' agregada"}
        return {"exito": False, "mensaje": "❌ Error al guardar categoría"}

    def eliminar_categoria(self, categoria: str) -> Dict:
        """Elimina una categoría personalizada."""
        if categoria not in self.categorias_personalizadas:
            return {"exito": False, "mensaje": f"❌ Categoría no encontrada"}
        
        self.categorias_personalizadas.remove(categoria)
        if self._guardar_categorias():
            return {"exito": True, "mensaje": f"✅ Categoría '{categoria}' eliminada"}
        return {"exito": False, "mensaje": "❌ Error al guardar"}

    def obtener_categorias(self) -> List[str]:
        """Retorna todas las categorías disponibles."""
        return self.categorias_personalizadas

    def establecer_presupuesto(self, presupuesto_diario: float = None, presupuesto_mensual: float = None) -> Dict:
        """Actualiza los límites de presupuesto."""
        try:
            if presupuesto_diario is not None:
                self.config["presupuesto_diario"] = presupuesto_diario
            if presupuesto_mensual is not None:
                self.config["presupuesto_mensual"] = presupuesto_mensual
            
            if self._guardar_config():
                return {
                    "exito": True,
                    "presupuesto_diario": self.config.get("presupuesto_diario"),
                    "presupuesto_mensual": self.config.get("presupuesto_mensual")
                }
            return {"exito": False, "mensaje": "❌ Error al guardar"}
        except Exception as e:
            logger.error(f"Error estableciendo presupuesto: {e}")
            return {"exito": False, "mensaje": f"❌ Error: {str(e)}"}

    def obtener_configuracion(self) -> Dict:
        """Retorna la configuración actual."""
        return self.config


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    manager = SpreadsheetManager(use_google_drive=False)
    print("SpreadsheetManager cargado correctamente")
